"""
Copyright (c) 2019 Yevheniia Soroka
Licensed under the MIT License
Author: Yevheniia Soroka
Email: ysoroka@cs.stonybrook.edu
Last modified: 07/08/2019

Usage:
Run this script to write accuracies of predictions for different 
datasets to Excel file -- from single dataset training.
"""

import os
import glob
import numpy as np
from openpyxl import *

ckpts_dir = os.getcwd().replace("PyUtils", "ckpts")

def process_log(log_path):
	'''
	For a single log file, finds test accuracies 
	where validation accuracies are the highest.
	'''
    vals = []
    test_precs = []
    pars = []

    f = open(log_path, "r")
    content = f.read().split("\n")
    for line in content:
        if "- freeze" in line:
            pars.append(str(line.split(" ")[-1][2:]))
        elif "- fix_BN" in line:
            pars.append(str(line.split(" ")[-1][2:]))
        else:
            val_split = line.split("- Validation * Prec@1 ")
            if len(val_split) > 1:
                vals.append(float(val_split[-1]))
            test_split = line.split("- Test * Prec@1 ")
            if len(test_split) > 1:
                test_precs.append(float(test_split[-1]))

    best_prec = test_precs[np.argmax(vals)]

    return best_prec #(dataset_name, pars, best_prec)


def write_to_xls(xlsx_file, log_data, row_nums, def_row):
	'''	Writes dataset accuracies to the Excel file. '''
    wb = load_workbook(xlsx_file)
    ws = wb['Sheet1']
    dataset_name, pars, prec = log_data

    if dataset_name not in row_nums.keys():
        row_nums[dataset_name] = def_row
    for col_no in range(1, 6):
        if ws.cell(column=col_no*2-1, row=1).value == dataset_name:
            ws.cell(column=col_no*2-1, row=row_nums[dataset_name]).value = '-'.join(pars)
            ws.cell(column=col_no*2, row=row_nums[dataset_name]).value = prec
            wb.save(xlsx_file)
            row_nums[dataset_name] += 1

    return row_nums

def main():
    row_nums = {}
    for file in glob.glob(ckpts_dir + "/SinglesFalse/Deepemotion/*.txt"):
        log_data = process_log(file)
        print(log_data)
        row_nums = write_to_xls("D:/Eugenia/Emotion/MetaEmotion/checkpts/comparison.xlsx", log_data, row_nums, def_row=9)


if __name__ == '__main__':
    main()